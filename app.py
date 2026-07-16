import csv
import hashlib
import os
from bisect import bisect_right
from io import BytesIO, StringIO

import chromadb
import streamlit as st
from docx import Document
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from pypdf import PdfReader
from sentence_transformers import SentenceTransformer


CHUNK_SIZE = 1000
CHUNK_OVERLAP = 200
TOP_K = 4
EMBEDDING_MODEL_NAME = "paraphrase-multilingual-MiniLM-L12-v2"
SIMILARITY_THRESHOLD = 0.25
CHROMA_DB_PATH = ".chroma_db"
SUPPORTED_FILE_TYPES = ["pdf", "txt", "md", "docx", "csv", "xlsx"]


load_dotenv()

st.set_page_config(page_title="企业知识库智能助手", layout="wide")
st.title("企业知识库智能助手")

st.session_state.setdefault("messages", [])
st.session_state.setdefault("active_knowledge_base", None)


def sha256_bytes(content):
    return hashlib.sha256(content).hexdigest()


def get_file_extension(file_name):
    return file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""


def make_file_set_id(file_infos):
    fingerprint_text = "|".join(
        f"{file_info['name']}:{file_info['hash']}" for file_info in file_infos
    )
    return hashlib.sha256(fingerprint_text.encode("utf-8")).hexdigest()


def safe_collection_name(file_set_id):
    return f"kb_{file_set_id[:24]}"


uploaded_files = st.file_uploader(
    "上传知识库文档",
    type=SUPPORTED_FILE_TYPES,
    accept_multiple_files=True,
)

file_infos = []
for uploaded_file in uploaded_files:
    file_bytes = uploaded_file.getvalue()
    file_infos.append(
        {
            "name": uploaded_file.name,
            "type": get_file_extension(uploaded_file.name),
            "bytes": file_bytes,
            "hash": sha256_bytes(file_bytes),
            "size": len(file_bytes),
        }
    )

if file_infos:
    file_set_id = make_file_set_id(file_infos)
    if st.session_state.active_knowledge_base != file_set_id:
        st.session_state.active_knowledge_base = file_set_id
        st.session_state.messages = []
else:
    file_set_id = None
    if st.session_state.active_knowledge_base is not None:
        st.session_state.active_knowledge_base = None
        st.session_state.messages = []

with st.form("question_form"):
    question = st.text_input("请输入你想问知识库的问题")
    submitted = st.form_submit_button("提交问题", type="primary")


@st.cache_data(max_entries=20)
def extract_pdf_units(file_name, file_bytes):
    reader = PdfReader(BytesIO(file_bytes))

    if reader.is_encrypted:
        raise ValueError("PDF is encrypted")

    units = []
    for page_number, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        lines = [line.strip() for line in page_text.splitlines() if line.strip()]
        if lines:
            units.append(
                {
                    "source_name": file_name,
                    "file_type": "pdf",
                    "location_type": "page_line",
                    "page": page_number,
                    "start_line": 1,
                    "end_line": len(lines),
                    "text": "\n".join(lines),
                }
            )

    return units


def decode_text(file_bytes):
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


@st.cache_data(max_entries=20)
def extract_text_units(file_name, file_bytes, file_type):
    text = decode_text(file_bytes)
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    if not lines:
        return []

    return [
        {
            "source_name": file_name,
            "file_type": file_type,
            "location_type": "line",
            "page": 0,
            "start_line": 1,
            "end_line": len(lines),
            "text": "\n".join(lines),
        }
    ]


@st.cache_data(max_entries=20)
def extract_docx_units(file_name, file_bytes):
    document = Document(BytesIO(file_bytes))
    units = []

    for paragraph_number, paragraph in enumerate(document.paragraphs, start=1):
        text = paragraph.text.strip()
        if text:
            units.append(
                {
                    "source_name": file_name,
                    "file_type": "docx",
                    "location_type": "paragraph",
                    "page": 0,
                    "start_line": paragraph_number,
                    "end_line": paragraph_number,
                    "text": text,
                }
            )

    return units


@st.cache_data(max_entries=20)
def extract_csv_units(file_name, file_bytes):
    text = decode_text(file_bytes)
    reader = csv.reader(StringIO(text))
    units = []

    for row_number, row in enumerate(reader, start=1):
        row_text = " | ".join(cell.strip() for cell in row if cell.strip())
        if row_text:
            units.append(
                {
                    "source_name": file_name,
                    "file_type": "csv",
                    "location_type": "row",
                    "page": 0,
                    "start_line": row_number,
                    "end_line": row_number,
                    "text": row_text,
                }
            )

    return units


@st.cache_data(max_entries=20)
def extract_xlsx_units(file_name, file_bytes):
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    units = []

    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                units.append(
                    {
                        "source_name": file_name,
                        "file_type": "xlsx",
                        "location_type": "sheet_row",
                        "sheet": worksheet.title,
                        "page": 0,
                        "start_line": row_number,
                        "end_line": row_number,
                        "text": f"{worksheet.title} | " + " | ".join(values),
                    }
                )

    return units


def extract_document_units(file_info):
    file_type = file_info["type"]

    if file_type == "pdf":
        return extract_pdf_units(file_info["name"], file_info["bytes"])
    if file_type in {"txt", "md"}:
        return extract_text_units(file_info["name"], file_info["bytes"], file_type)
    if file_type == "docx":
        return extract_docx_units(file_info["name"], file_info["bytes"])
    if file_type == "csv":
        return extract_csv_units(file_info["name"], file_info["bytes"])
    if file_type == "xlsx":
        return extract_xlsx_units(file_info["name"], file_info["bytes"])

    return []


def format_location(metadata):
    location_type = metadata.get("location_type")

    if location_type == "page_line":
        if metadata["start_line"] == metadata["end_line"]:
            return f"第 {metadata['page']} 页，第 {metadata['start_line']} 行"
        return f"第 {metadata['page']} 页，第 {metadata['start_line']}-{metadata['end_line']} 行"

    if location_type == "paragraph":
        return f"第 {metadata['start_line']} 段"

    if location_type == "row":
        return f"第 {metadata['start_line']} 行"

    if location_type == "sheet_row":
        return f"{metadata.get('sheet', '工作表')}，第 {metadata['start_line']} 行"

    if location_type == "line":
        return f"第 {metadata['start_line']}-{metadata['end_line']} 行"

    return "未知位置"


def split_units_to_chunks(units, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    chunks = []

    for unit in units:
        text = unit["text"]
        if not text:
            continue

        line_offsets = []
        position = 0
        lines = text.splitlines() or [text]
        for line in lines:
            line_offsets.append(position)
            position += len(line) + 1

        start = 0
        while start < len(text):
            end = min(start + chunk_size, len(text))
            raw_chunk_text = text[start:end]
            chunk_text = raw_chunk_text.strip()

            if chunk_text:
                leading_space = len(raw_chunk_text) - len(raw_chunk_text.lstrip())
                trailing_space = len(raw_chunk_text) - len(raw_chunk_text.rstrip())
                content_start = start + leading_space
                content_end = end - trailing_space

                start_line = unit["start_line"]
                end_line = unit["end_line"]
                if unit["location_type"] in {"page_line", "line"}:
                    start_line = unit["start_line"] + bisect_right(line_offsets, content_start) - 1
                    end_line = unit["start_line"] + bisect_right(line_offsets, content_end - 1) - 1

                metadata = {
                    "source_name": unit["source_name"],
                    "file_type": unit["file_type"],
                    "location_type": unit["location_type"],
                    "page": int(unit.get("page", 0)),
                    "start_line": int(start_line),
                    "end_line": int(end_line),
                    "sheet": unit.get("sheet", ""),
                    "chunk_id": len(chunks) + 1,
                }

                chunks.append(
                    {
                        "id": str(len(chunks) + 1),
                        "text": chunk_text,
                        "metadata": metadata,
                    }
                )

            if end == len(text):
                break

            start = end - overlap

    return chunks


def build_preview_text(chunks, limit=5000):
    sections = []
    current_length = 0

    for chunk in chunks:
        metadata = chunk["metadata"]
        location = format_location(metadata)
        section = (
            f"--- {metadata['source_name']} | {location} | 文本块 {metadata['chunk_id']} ---\n"
            f"{chunk['text']}"
        )
        if current_length + len(section) > limit:
            break
        sections.append(section)
        current_length += len(section)

    return "\n\n".join(sections)


@st.cache_resource
def load_embedding_model():
    return SentenceTransformer(EMBEDDING_MODEL_NAME)


@st.cache_resource
def get_chroma_client():
    return chromadb.PersistentClient(path=CHROMA_DB_PATH)


def get_chroma_collection(collection_name):
    client = get_chroma_client()
    return client.get_or_create_collection(
        name=collection_name,
        metadata={"hnsw:space": "cosine"},
    )


@st.cache_data(max_entries=10, show_spinner=False)
def build_embeddings(chunk_texts):
    model = load_embedding_model()
    return model.encode(
        list(chunk_texts),
        normalize_embeddings=True,
        show_progress_bar=False,
    ).tolist()


def index_chunks_in_chroma(collection_name, chunks):
    collection = get_chroma_collection(collection_name)

    if collection.count() == len(chunks):
        return collection

    if collection.count() > 0:
        existing = collection.get()
        if existing.get("ids"):
            collection.delete(ids=existing["ids"])

    chunk_texts = tuple(chunk["text"] for chunk in chunks)
    embeddings = build_embeddings(chunk_texts)

    collection.add(
        ids=[chunk["id"] for chunk in chunks],
        documents=[chunk["text"] for chunk in chunks],
        metadatas=[chunk["metadata"] for chunk in chunks],
        embeddings=embeddings,
    )

    return collection


def retrieve_relevant_chunks(collection, question, top_k=TOP_K):
    model = load_embedding_model()
    question_embedding = model.encode(
        [question],
        normalize_embeddings=True,
        show_progress_bar=False,
    ).tolist()[0]

    results = collection.query(
        query_embeddings=[question_embedding],
        n_results=top_k,
        include=["documents", "metadatas", "distances"],
    )

    relevant_chunks = []
    documents = results.get("documents", [[]])[0]
    metadatas = results.get("metadatas", [[]])[0]
    distances = results.get("distances", [[]])[0]

    for document, metadata, distance in zip(documents, metadatas, distances):
        score = round(1 - float(distance), 3)
        if score >= SIMILARITY_THRESHOLD:
            relevant_chunks.append(
                {
                    "text": document,
                    "metadata": metadata,
                    "score": score,
                }
            )

    return relevant_chunks


def build_context(chunks):
    context_sections = []

    for chunk in chunks:
        metadata = chunk["metadata"]
        location = format_location(metadata)
        context_sections.append(
            (
                f"[来源：{metadata['source_name']}，{location}，"
                f"文本块 {metadata['chunk_id']}]\n{chunk['text']}"
            )
        )

    return "\n\n".join(context_sections)


def get_deepseek_api_key():
    api_key = os.getenv("DEEPSEEK_API_KEY")

    if api_key:
        return api_key

    try:
        return st.secrets["DEEPSEEK_API_KEY"]
    except Exception:
        return None


def ask_ai(context_text, user_question):
    api_key = get_deepseek_api_key()

    if not api_key:
        return "还没有配置 DEEPSEEK_API_KEY。你可以先完成知识库上传和检索，下一步再接入 AI。"

    client = OpenAI(
        api_key=api_key,
        base_url="https://api.deepseek.com",
    )

    prompt = f"""
你是一个严谨的企业知识库问答助手。
请只根据下面通过 Chroma 向量数据库检索到的知识库文本块回答用户问题。
如果这些文本块中没有答案，请说：文档中没有找到相关信息。
回答关键信息时，请在对应句子末尾标注来源，格式为：[文档名，第 X 页/第 X 行/第 X 段]。
只能使用文本块中提供的来源信息，不要编造来源。

检索到的知识库文本块：
{context_text[:12000]}

用户问题：
{user_question}
"""

    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
        )

        return response.choices[0].message.content

    except Exception as e:
        error_text = str(e).lower()

        if "402" in error_text or "insufficient balance" in error_text or "balance" in error_text:
            return "当前 DeepSeek 账户余额不足，请充值后重试，或更换可用的 API Key。"

        if "401" in error_text or "authentication" in error_text or "api key" in error_text:
            return "DeepSeek API Key 无效或未授权，请检查 .env 文件中的 DEEPSEEK_API_KEY。"

        if "429" in error_text or "rate limit" in error_text:
            return "当前请求过于频繁，触发了 DeepSeek 限流，请稍后再试。"

        if "500" in error_text or "503" in error_text or "server" in error_text or "overloaded" in error_text:
            return "DeepSeek 服务暂时繁忙，请稍后再试。"

        if "timeout" in error_text or "connection" in error_text:
            return "连接 DeepSeek 服务失败，请检查网络后重试。"

        return "AI 服务暂时不可用，请稍后再试。"


if submitted and not file_infos:
    st.warning("请先上传至少一个知识库文档。")

if file_infos:
    try:
        with st.spinner("正在解析知识库文档..."):
            document_units = []
            for file_info in file_infos:
                document_units.extend(extract_document_units(file_info))

            text_chunks = split_units_to_chunks(document_units)

    except Exception as e:
        error_text = str(e).lower()

        if "encrypt" in error_text or "password" in error_text:
            st.error("文档读取失败：其中一个 PDF 已加密，请解除密码保护后再上传。")
        else:
            st.error("文档读取失败，请确认文件没有损坏，并且格式受支持。")

        st.stop()

    if not text_chunks:
        st.warning("没有从文档中提取到可检索文本。扫描版 PDF 后续需要加入 OCR。")
    else:
        collection_name = safe_collection_name(file_set_id)

        with st.spinner("正在写入 Chroma 向量数据库..."):
            collection = index_chunks_in_chroma(collection_name, text_chunks)

        st.success(
            f"知识库构建成功：{len(file_infos)} 个文档，"
            f"{len(text_chunks)} 个文本块，已写入 Chroma。"
        )
        st.caption("PDF 行号基于提取后的文本计算，可能与复杂排版中的视觉行号略有差异。")

        with st.expander("查看已上传文档"):
            st.table(
                [
                    {
                        "文档名": file_info["name"],
                        "格式": file_info["type"],
                        "大小 KB": round(file_info["size"] / 1024, 1),
                    }
                    for file_info in file_infos
                ]
            )

        with st.expander("查看提取到的知识库文本"):
            st.text_area("知识库文本预览", build_preview_text(text_chunks), height=300)

        if submitted:
            if not question.strip():
                st.warning("请输入问题后再提交。")
            else:
                question_text = question.strip()
                relevant_chunks = retrieve_relevant_chunks(collection, question_text)
                st.session_state.messages.append(
                    {"role": "user", "content": question_text}
                )

                if relevant_chunks:
                    context_text = build_context(relevant_chunks)
                    with st.spinner("AI 正在思考..."):
                        answer = ask_ai(context_text, question_text)
                else:
                    answer = (
                        "没有在知识库中检索到与问题相关的内容，"
                        "因此本次没有调用 AI。请换一种问法，或确认文档中是否包含相关信息。"
                    )

                st.session_state.messages.append(
                    {
                        "role": "assistant",
                        "content": answer,
                        "sources": relevant_chunks,
                    }
                )

        if st.session_state.messages:
            st.subheader("聊天记录")

            chat_rounds = [
                st.session_state.messages[index:index + 2]
                for index in range(0, len(st.session_state.messages), 2)
            ]

            for chat_round in reversed(chat_rounds):
                for message in chat_round:
                    with st.chat_message(message["role"]):
                        st.write(message["content"])

                        sources = message.get("sources", [])
                        if sources:
                            with st.expander(f"查看检索到的 {len(sources)} 个文本块"):
                                for source in sources:
                                    metadata = source["metadata"]
                                    source_location = format_location(metadata)

                                    st.markdown(
                                        f"**{metadata['source_name']}，{source_location}，"
                                        f"文本块 {metadata['chunk_id']}，"
                                        f"相似度 {source['score']}**"
                                    )
                                    st.write(source["text"])

            if st.button("清空聊天记录"):
                st.session_state.messages = []
                st.rerun()
