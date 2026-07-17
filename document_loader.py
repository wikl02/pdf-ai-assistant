import csv
import hashlib
from functools import lru_cache
from io import BytesIO, StringIO

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def sha256_bytes(content):
    return hashlib.sha256(content).hexdigest()


def get_file_extension(file_name):
    return file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""


def make_file_set_id(file_infos):
    fingerprint_text = "|".join(
        f"{file_info['name']}:{file_info['hash']}" for file_info in file_infos
    )
    return hashlib.sha256(fingerprint_text.encode("utf-8")).hexdigest()


def safe_collection_name(file_set_id):
    return f"kb_{file_set_id[:24]}"


def build_file_infos(uploaded_files):
    file_infos = []

    for uploaded_file in uploaded_files:
        file_bytes = uploaded_file.getvalue()
        file_infos.append(
            {
                "name": uploaded_file.name,
                "type": get_file_extension(uploaded_file.name),
                "bytes": file_bytes,
                "hash": sha256_bytes(file_bytes),
                "size": len(file_bytes),
            }
        )

    return file_infos


def decode_text(file_bytes):
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


@lru_cache(maxsize=20)
def extract_pdf_units(file_name, file_bytes):
    reader = PdfReader(BytesIO(file_bytes))

    if reader.is_encrypted:
        raise ValueError("PDF is encrypted")

    units = []
    for page_number, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        lines = [line.strip() for line in page_text.splitlines() if line.strip()]
        if lines:
            units.append(
                {
                    "source_name": file_name,
                    "file_type": "pdf",
                    "location_type": "page_line",
                    "page": page_number,
                    "start_line": 1,
                    "end_line": len(lines),
                    "text": "\n".join(lines),
                }
            )

    return units


@lru_cache(maxsize=20)
def extract_text_units(file_name, file_bytes, file_type):
    text = decode_text(file_bytes)
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    if not lines:
        return []

    return [
        {
            "source_name": file_name,
            "file_type": file_type,
            "location_type": "line",
            "page": 0,
            "start_line": 1,
            "end_line": len(lines),
            "text": "\n".join(lines),
        }
    ]


@lru_cache(maxsize=20)
def extract_docx_units(file_name, file_bytes):
    document = Document(BytesIO(file_bytes))
    units = []

    for paragraph_number, paragraph in enumerate(document.paragraphs, start=1):
        text = paragraph.text.strip()
        if text:
            units.append(
                {
                    "source_name": file_name,
                    "file_type": "docx",
                    "location_type": "paragraph",
                    "page": 0,
                    "start_line": paragraph_number,
                    "end_line": paragraph_number,
                    "text": text,
                }
            )

    return units


@lru_cache(maxsize=20)
def extract_csv_units(file_name, file_bytes):
    text = decode_text(file_bytes)
    reader = csv.reader(StringIO(text))
    units = []

    for row_number, row in enumerate(reader, start=1):
        row_text = " | ".join(cell.strip() for cell in row if cell.strip())
        if row_text:
            units.append(
                {
                    "source_name": file_name,
                    "file_type": "csv",
                    "location_type": "row",
                    "page": 0,
                    "start_line": row_number,
                    "end_line": row_number,
                    "text": row_text,
                }
            )

    return units


@lru_cache(maxsize=20)
def extract_xlsx_units(file_name, file_bytes):
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    units = []

    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [
                str(value).strip()
                for value in row
                if value is not None and str(value).strip()
            ]
            if values:
                units.append(
                    {
                        "source_name": file_name,
                        "file_type": "xlsx",
                        "location_type": "sheet_row",
                        "sheet": worksheet.title,
                        "page": 0,
                        "start_line": row_number,
                        "end_line": row_number,
                        "text": f"{worksheet.title} | " + " | ".join(values),
                    }
                )

    return units


def extract_document_units(file_info):
    file_type = file_info["type"]

    if file_type == "pdf":
        return extract_pdf_units(file_info["name"], file_info["bytes"])
    if file_type in {"txt", "md"}:
        return extract_text_units(file_info["name"], file_info["bytes"], file_type)
    if file_type == "docx":
        return extract_docx_units(file_info["name"], file_info["bytes"])
    if file_type == "csv":
        return extract_csv_units(file_info["name"], file_info["bytes"])
    if file_type == "xlsx":
        return extract_xlsx_units(file_info["name"], file_info["bytes"])

    return []
